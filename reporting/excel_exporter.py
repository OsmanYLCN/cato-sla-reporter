from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config.settings import (
    COLOR_FAILED_BG,
    COLOR_FAILED_FONT,
    COLOR_HEADER_BG,
    COLOR_HEADER_FONT,
    COLOR_PASSED_BG,
    COLOR_PASSED_FONT,
    COLOR_ROW_ALT,
    EXCEL_SHEET_DETAILS,
    EXCEL_SHEET_OVERALL,
    EXCEL_SHEET_SUMMARY,
    OUTPUT_DIR,
)
from engine.sla_calculator import (
    COL_OUT_AVAIL,
    COL_OUT_COUNT,
    COL_OUT_DURATION,
    COL_OUT_PERIOD,
    COL_OUT_SITE,
    COL_OUT_SLA,
)
from engine.state_machine import OutageRecord
from utils.logger import get_logger

logger = get_logger(__name__)

# Excel tablo stilleri
_THIN_SIDE = Side(style="thin", color="FFB8CCE4")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_HEADER_FILL = PatternFill("solid", fgColor=COLOR_HEADER_BG)
_ALT_ROW_FILL = PatternFill("solid", fgColor=COLOR_ROW_ALT)
_PASSED_FILL = PatternFill("solid", fgColor=COLOR_PASSED_BG)
_PASSED_FONT_STYLE = Font(bold=True, color=COLOR_PASSED_FONT)
_FAILED_FILL = PatternFill("solid", fgColor=COLOR_FAILED_BG)
_FAILED_FONT_STYLE = Font(bold=True, color=COLOR_FAILED_FONT)
_HEADER_FONT = Font(bold=True, color=COLOR_HEADER_FONT, name="Calibri", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT = Alignment(horizontal="left", vertical="center")

# Excel raporu dışa aktarma
def export_to_excel(
    summary_df: pd.DataFrame,
    outages: list[OutageRecord],
    period_months: int,
    output_dir: str | Path | None = None,
    report_date: date | None = None,
) -> Path:
    """SLA ozet tablosunu, kesinti detaylarini ve genel ortalama ozetini
    uc sekme halinde Excel dosyasina aktarir."""
    out_dir = Path(output_dir) if output_dir else Path(OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    report_date = report_date or date.today()
    filename = f"SLA_Report_{period_months}M_{report_date.strftime('%Y-%m-%d')}_{datetime.now().strftime('%H-%M-%S')}.xlsx"
    file_path = out_dir / filename

    logger.info("Excel raporu olusturuluyor: %s", file_path)

    wb = Workbook()

    # Sekme 1: SLA Ozet tablosu
    ws_summary = wb.active
    ws_summary.title = EXCEL_SHEET_SUMMARY
    _build_summary_sheet(ws_summary, summary_df)

    # Sekme 2: Ham kesinti detaylari
    ws_details = wb.create_sheet(title=EXCEL_SHEET_DETAILS)
    _build_details_sheet(ws_details, outages)

    # Sekme 3: Overall Summary — tum sitelerin genel ortalamasi (Ingilizce)
    ws_overall = wb.create_sheet(title=EXCEL_SHEET_OVERALL)
    _build_overall_sheet(ws_overall, summary_df, period_months)

    try:
        wb.save(file_path)
        logger.info("Excel raporu basariyla yazildi: %s", file_path)
    except OSError as exc:
        raise OSError(
            f"Excel dosyasi yazilamadi '{file_path}': {exc}"
        ) from exc

    return file_path


def _sanitize_cell_value(value: any) -> any:
    """Excel Formula Injection (CWE-1236) koruması sağlar.
    Metin '=', '+', '-', '@' ile başlıyorsa Excel'in formül olarak çalıştırmasını önler.
    """
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


# SLA özet sekmesi
def _build_summary_sheet(ws, df: pd.DataFrame) -> None:
    """SLA Özet sekmesini biçimlendirir ve verileri ekler."""
    headers = [
        COL_OUT_SITE,
        COL_OUT_PERIOD,
        COL_OUT_COUNT,
        COL_OUT_DURATION,
        COL_OUT_AVAIL,
        COL_OUT_SLA,
    ]

    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    ws.row_dimensions[1].height = 22

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        sla_status = row[COL_OUT_SLA]
        is_alt_row = (row_idx % 2 == 0)

        row_data = [
            _sanitize_cell_value(row[COL_OUT_SITE]),
            _sanitize_cell_value(row[COL_OUT_PERIOD]),
            int(row[COL_OUT_COUNT]),
            row[COL_OUT_DURATION],
            row[COL_OUT_AVAIL],
            _sanitize_cell_value(sla_status),
        ]

        ws.append(row_data)

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER if col_idx != 1 else _LEFT

            if is_alt_row:
                cell.fill = _ALT_ROW_FILL

            if col_idx == 6:
                if sla_status == "Passed":
                    cell.fill = _PASSED_FILL
                    cell.font = _PASSED_FONT_STYLE
                else:
                    cell.fill = _FAILED_FILL
                    cell.font = _FAILED_FONT_STYLE

            if col_idx == 4:
                cell.number_format = "0.00"
            elif col_idx == 5:
                cell.number_format = '0.0000"%"'

        ws.row_dimensions[row_idx].height = 18

    col_widths = [32, 16, 22, 30, 20, 16]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Tablo üst başlık
    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "CATO NETWORKS — SLA / AVAILABILITY REPORT"
    title_cell.font = Font(bold=True, size=14, color="FFFFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="FF1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.freeze_panes = "A3"

# Kesinti Detayları sekmesi
def _build_details_sheet(ws, outages: list[OutageRecord]) -> None:
    """Kesinti Detayları sekmesini oluşturur ve zaman sıralı kesintileri listeler."""
    headers = ["Site Name", "Start Time", "End Time", "Duration (Minutes)"]
    date_fmt = "%d.%m.%Y %H:%M:%S"

    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 22

    if not outages:
        ws.append(["No outages detected during the report period.", "", "", ""])
        logger.info("Kesinti detayları sekmesi: kesinti yok.")
        return

    sorted_outages = sorted(outages, key=lambda o: (o.site, o.start))

    for row_idx, rec in enumerate(sorted_outages, start=2):
        is_alt_row = (row_idx % 2 == 0)
        row_data = [
            _sanitize_cell_value(rec.site),
            rec.start.strftime(date_fmt),
            rec.end.strftime(date_fmt),
            rec.duration_minutes,
        ]
        ws.append(row_data)

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER if col_idx != 1 else _LEFT
            if is_alt_row:
                cell.fill = _ALT_ROW_FILL
            if col_idx == 4:
                cell.number_format = "0.00"

        ws.row_dimensions[row_idx].height = 17

    col_widths = [32, 22, 22, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    logger.info("Kesinti Detaylari sekmesi yazildi: %d kayit.", len(sorted_outages))


# Overall Summary sekmesi
def _build_overall_sheet(
    ws,
    df: pd.DataFrame,
    period_months: int,
) -> None:
    """Overall Summary sekmesini olusturur.

    Tum sitelerin rapor donemindeki metriklerinin ortalamasini tek bir
    ozet tablo seklinde gosterir. Tum etiketler ve degerler Ingilizce'dir.
    """
    from config.settings import SLA_THRESHOLD_PCT

    # --- Baslik ---
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "CATO NETWORKS — OVERALL AVAILABILITY SUMMARY"
    title_cell.font = Font(bold=True, size=14, color="FFFFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="FF1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # --- Bos satir ---
    ws.append(["", ""])

    def _section_row(label: str, value: str | int | float, number_fmt: str | None = None) -> None:
        """Etiket-deger cifti satirini yazar ve stilini uygular."""
        ws.append([label, value])
        row_idx = ws.max_row
        label_cell = ws.cell(row=row_idx, column=1)
        value_cell = ws.cell(row=row_idx, column=2)

        label_cell.font  = Font(bold=True, name="Calibri", size=11)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = _THIN_BORDER

        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = _THIN_BORDER
        if number_fmt:
            value_cell.number_format = number_fmt

        ws.row_dimensions[row_idx].height = 18

    # --- Rapor Bilgisi ---
    period_label = f"{period_months} Month" if period_months == 1 else f"{period_months} Months"
    _section_row("Report Period",    period_label)
    _section_row("Total Sites",      len(df) if not df.empty else 0)
    _section_row("SLA Threshold",    f"{SLA_THRESHOLD_PCT:.2f}%")

    ws.append(["", ""])  # bos satir

    # --- Ortalama Metrikler (tum siteler) ---
    if not df.empty:
        avg_avail     = df[COL_OUT_AVAIL].mean()
        avg_downtime  = df[COL_OUT_DURATION].mean()
        avg_outages   = df[COL_OUT_COUNT].mean()
        total_outages = int(df[COL_OUT_COUNT].sum())
        passed_count  = int((df[COL_OUT_SLA] == "Passed").sum())
        failed_count  = int((df[COL_OUT_SLA] == "Failed").sum())
        pass_rate     = (passed_count / len(df)) * 100 if len(df) > 0 else 0.0
    else:
        avg_avail = avg_downtime = avg_outages = pass_rate = 0.0
        total_outages = passed_count = failed_count = 0

    _section_row("Avg. Availability (%)",          round(avg_avail, 4),    '0.0000"%"')
    _section_row("Avg. Downtime (Minutes)",         round(avg_downtime, 2), "0.00")
    _section_row("Avg. Outage Count per Site",      round(avg_outages, 2),  "0.00")
    _section_row("Total Outages (All Sites)",        total_outages)

    ws.append(["", ""])

    _section_row("Sites Passed SLA",   passed_count)
    _section_row("Sites Failed SLA",   failed_count)
    _section_row("SLA Pass Rate (%)",  round(pass_rate, 2), '0.00"%"')

    # --- Pass Rate hucresini renkledir ---
    pass_rate_row = ws.max_row
    cell = ws.cell(row=pass_rate_row, column=2)
    if pass_rate >= 100.0:
        cell.fill = _PASSED_FILL
        cell.font = _PASSED_FONT_STYLE
    elif pass_rate < 75.0:
        cell.fill = _FAILED_FILL
        cell.font = _FAILED_FONT_STYLE

    # --- Sutun genislikleri ---
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    logger.info("Overall Summary sekmesi yazildi.")
